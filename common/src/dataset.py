
import tempfile
import os.path
import sys
import re
import itertools
import traceback
import copy
import gzip
import bz2
import warnings
import zipfile
import shutil
from collections import defaultdict
from operator import itemgetter
try:
    import pickle as Pickle
except ImportError:
    import Pickle


class Dataset(object):

    def __init__(self):
        self.tables = []
        self.namemap = {}
        self.current = None

    def new_table(self, name, table, rows):
        self.current = len(self.tables)
        self.tables.append(table)
        self.namemap[name] = self.current
        table.from_rows(rows)


class FileDataset(Dataset):

    def __init__(self, filename=None, names=None, tables=None):
        if not filename:
            filedesc, self.filename = tempfile.mkstemp()
            os.close(filedesc)
            self.tempfile = True
            assert names and tables
            self.names_ = names
            self.from_tables(tables)
        else:
            self.filename = filename
            self.tempfile = False
            if not names and not tables:
                self.set_names()
            else:
                self.names_ = names
                self.from_tables(tables)

class XLSFileDataset(FileDataset):

    def __init__(self, *args, **kw):
        import xlwt
        self.colwidth = defaultdict(lambda: {})
        if 'columnwidth' in kw:
            self.colwidth.update(kw['columnwidth'])
            del kw['columnwidth']
        self.rowheight = defaultdict(lambda: {})
        if 'rowheight' in kw:
            self.rowheight.update(kw['rowheight'])
            del kw['rowheight']
        self.underline = xlwt.easyxf('font: underline single')
        super(XLSFileDataset, self).__init__(*args, **kw)
    heightstyles = {}

    @staticmethod
    def setcolwidth(ws, j, pixels):
        winpts = int(pixels * 256.0 / 7.0)
        if ws.col(j).width < winpts:
            ws.col(j).width = winpts

    @staticmethod
    def setrowheight(ws, i, pixels):
        import xlwt
        height = XLSFileDataset.heightstyles.get(pixels)
        if height == None:
            hinpts = int(pixels * 12.0)
            height = xlwt.easyxf('font:height %d;' % hinpts)
            XLSFileDataset.heightstyles[pixels] = height
        ws.row(i).set_style(height)

    @staticmethod
    def writevalue(ws, i, j, value):
        import xlwt
        if value == None:
            return
        try:
            value = Pickle.loads(value)
        except (Pickle.UnpicklingError, ValueError, EOFError, TypeError):
            pass
        if isinstance(value, dict):
            if value['type'] == 'url':
                ws.write(i, j,
                         xlwt.Formula('HYPERLINK("%s";"%s")' %
                                      (value['url'], value['text'])),
                         XLSFileDataset.underline)
                return
            elif value['type'] == 'image':
                ws.insert_bitmap(value['image'], i, j)
                return
        try:
            ws.write(i, j, float(value))
            return
        except (ValueError, TypeError):
            pass
        try:
            ws.write(i, j, int(value))
            return
        except (ValueError, TypeError):
            pass
        ws.write(i, j, str(value))

    def set_names(self):
        import xlrd
        book = xlrd.open_workbook(self.filename)
        for sheet_idx, sheet_name in enumerate(book.sheet_names()):
            self.names_.append(sheet_name)

    def from_tables(self, tables):
        import xlwt
        wb = xlwt.Workbook()
        for n, t in zip(self.names_, tables):
            ws = wb.add_sheet(n)
            dcw = self.colwidth[n].get(None)
            for j, h in enumerate(t.headers()):
                cw = self.colwidth[n].get(h, dcw)
                if cw != None:
                    XLSFileDataset.setcolwidth(ws, j, cw)
            for j, h in enumerate(t.headers()):
                XLSFileDataset.writevalue(ws, 0, j, h)
            drh = self.rowheight[n].get(None)
            for i, r in enumerate(t.rows()):
                if i >= 65535:
                    break
                rh = self.rowheight[n].get(i, drh)
                if rh:
                    XLSFileDataset.setrowheight(ws, i + 1, rh)
                for j, h in enumerate(t.headers()):
                    XLSFileDataset.writevalue(ws, i + 1, j, r[h])
        wb.save(self.filename)

class XLSXFileDataset(FileDataset):

    def __init__(self, *args, **kw):
        self.colwidth = defaultdict(lambda: {})
        if 'columnwidth' in kw:
            self.colwidth.update(kw['columnwidth'])
            del kw['columnwidth']
        self.rowheight = defaultdict(lambda: {})
        if 'rowheight' in kw:
            self.rowheight.update(kw['rowheight'])
            del kw['rowheight']
        super(XLSXFileDataset, self).__init__(*args, **kw)

    @staticmethod
    def setcolwidth(ws, j, pixels):
        import openpyxl
        # openpyxl.shared.units.pixels_to_points(pixels)
        winpts = pixels * 0.14214
        ws.cell(row=1, column=j+1)
        ws.column_dimensions[
            openpyxl.utils.cell.get_column_letter(j + 1)].width = winpts

    @staticmethod
    def setrowheight(ws, i, pixels):
        hinpts = pixels * 3. / \
            4.  # openpyxl.shared.units.pixels_to_points(pixels);
        ws.cell(row=i+1, column=1)
        ws.row_dimensions[i + 1].height = hinpts

    @staticmethod
    def writevalue(ws, i, j, value):
        import openpyxl
        if value == None:
            return
        try:
            value = Pickle.loads(value)
        except (Pickle.UnpicklingError, ValueError, EOFError, TypeError):
            pass
        if isinstance(value, dict):
            if value['type'] == 'url':
                ws.cell(row=i, column=j).value = value['text']
                ws.cell(row=i, column=j).hyperlink = value['url']
                ws.cell(row=i, column=j).style.font.underline = openpyxl.style.Font.UNDERLINE_SINGLE
                return
            if value['type'] == 'image':
                img = openpyxl.drawing.Image(value['image'])
                img.anchor(ws.cell(row=i, column=j),
                           type=openpyxl.drawing.Image.ONECELL)
                ws.add_image(img)
                return
        try:
            if float('-inf') < float(value) < float('+inf'):
                ws.cell(row=i, column=j).value = float(value)
            else:
                raise ValueError()
            return
        except (ValueError, TypeError):
            pass
        try:
            ws.cell(row=i, column=j).value = int(value)
            return
        except (ValueError, TypeError):
            pass
        ws.cell(row=i, column=j).value = str(value)

    def set_names(self):
        import openpyxl
        book = openpyxl.reader.excel.load_workbook(filename=self.filename)
        for sheet_idx, sheet_name in enumerate(book.get_sheet_names()):
            self.names_.append(sheet_name)

    def from_tables(self, tables):
        import openpyxl
        wb = openpyxl.workbook.Workbook()
        for i, (n, t) in enumerate(zip(self.names_, tables)):
            if i == 0:
                ws = wb.worksheets[0]
            else:
                ws = wb.create_sheet()
            ws.title = n
            # ws.auto_filter = 'A1:%s1' % (
            #     openpyxl.utils.cell.get_column_letter(len(t.headers())),)
            dcw = self.colwidth[n].get(None)
            for j, h in enumerate(t.headers()):
                cw = self.colwidth[n].get(h, dcw)
                if cw != None:
                    XLSXFileDataset.setcolwidth(ws, j, cw)
            for j, h in enumerate(t.headers()):
                ws.cell(row=1, column=j+1).value = h
                ws.cell(row=1, column=j+1).style.font.bold = True
                ws.cell(row=1, column=j+1).style.borders.bottom.border_style = openpyxl.style.Border.BORDER_THIN
            drh = self.rowheight[n].get(None)
            for i, r in enumerate(t.rows()):
                rh = self.rowheight[n].get(i, drh)
                if rh:
                    XLSXFileDataset.setrowheight(ws, i + 1, rh)
                for j, h in enumerate(t.headers()):
                    XLSXFileDataset.writevalue(ws, i + 1, j, r.get(h))
        wb.save(self.filename)


class ZIPFileDataset(FileDataset):

    def from_tables(self, tables):
        tmpdir = tempfile.mkdtemp()
        for n, t in zip(self.names_, tables):
            to = CSVFileTable(filename=os.path.join(tmpdir, n + '.csv'),
                              from_rows=t, compression=None)
        zf = zipfile.ZipFile(self.filename, 'w',
                             compression=zipfile.ZIP_DEFLATED, allowZip64=True)
        for n in self.names_:
            zf.write(os.path.join(tmpdir, n + '.csv'), n + '.csv')
        zf.close()
        shutil.rmtree(tmpdir, ignore_errors=True)


class Table(object):

    def extract(self, *keys):
        for r in self.rows():
            if len(keys) == 1:
                yield r[keys[0]]
            else:
                yield [r[k] for k in keys]

    def headers(self):
        return self.headers_

    def __iter__(self):
        return self.rows()


class MemoryTable(Table):

    def __init__(self, headers):
        self.headers_ = list(headers)
        self.therows = []

    def rows(self):
        for r in self.therows:
            yield dict(list(zip(self.headers_, r)))

    def size(self):
        return len(self.therows)

    def from_rows(self, rows):
        self.therows = []
        for r in rows:
            self.therows.append([r.get(h, "") for h in self.headers_])

    def sort(self, key=None, cmp=None):
        if key:
            self.therows.sort(key=lambda r: key(dict(list(zip(self.headers_, r)))))
        else:
            self.therows.sort(cmp=lambda a, b: cmp(
                dict(list(zip(self.headers_, a))), dict(list(zip(self.headers_, b)))))


class FileTable(Table):

    def __init__(self, filename=None, headers=None, from_rows=None, compression=None):
        self.compression = compression
        if not filename:
            cmp = ''
            if compression:
                cmp = '.' + compression
            filedesc, self.filename = tempfile.mkstemp(suffix=cmp)
            os.close(filedesc)
            self.tempfile = True
            if from_rows:
                self.headers_ = from_rows.headers()
                self.from_rows(from_rows)
            else:
                assert(headers)
                self.headers_ = headers
        else:
            self.filename = filename
            self.tempfile = False
            if not headers and not from_rows:
                self.set_headers()
            else:
                if from_rows:
                    self.headers_ = from_rows.headers()
                    self.from_rows(from_rows)
                else:
                    self.headers_ = headers

    def __del__(self):
        if hasattr(self, 'tempfile') and self.tempfile:
            if os.path.exists(self.filename):
                os.unlink(self.filename)

    def open(self, mode='rt', encoding='utf8'):
        if not isinstance(self.filename, str):
            return iter(self.filename)
        if self.filename.lower().endswith(".gz"):
            if 'r' in mode:
                h = gzip.open(self.filename, mode=mode, encoding=encoding)
                h.readline()
                h.seek(0)
            else:
                h = gzip.open(self.filename, mode=mode, encoding=encoding)
            return h
        elif self.filename.lower().endswith(".bz2"):
            if 'r' in mode:
                h = bz2.BZ2File(self.filename, mode=mode, encoding=encoding)
                h.readline()
                h.seek(0)
            else:
                h = bz2.BZ2File(self.filename, mode=mode, encoding=encoding)
            return h
        return open(self.filename, mode=mode, encoding=encoding)

import csv


class CSVFileTable(FileTable):

    def set_headers(self):
        h = self.open()
        t = csv.reader(h)
        self.headers_ = next(t)
        h.close()

    def rows(self):
        h = self.open()
        rows = csv.DictReader(h)
        for r in rows:
            yield r
        h.close()

    def from_rows(self, rows):
        wh = self.open(mode='w')
        dwh = csv.DictWriter(wh, self.headers_, extrasaction='ignore')
        dwh.writerow(dict([(h, h) for h in self.headers_]))
        for r in rows:
            dwh.writerow(r)
        wh.close()


class TSVFileTable(FileTable):

    def set_headers(self):
        h = self.open()
        t = csv.reader(h, dialect='excel-tab')
        self.headers_ = next(t)
        h.close()

    def rows(self):
        h = self.open()
        rows = csv.DictReader(h, dialect='excel-tab')
        for r in rows:
            yield r
        h.close()

    def from_rows(self, rows):
        wh = self.open(mode='w')
        dwh = csv.DictWriter(
            wh, self.headers_, extrasaction='ignore', dialect='excel-tab')
        dwh.writerow(dict([(h, h) for h in self.headers_]))
        for r in rows:
            dwh.writerow(r)
        wh.close()


class TXTFileTable(FileTable):

    def __init__(self, *args, **kw):
        assert('headers' in kw)
        self.delimeter = None
        if 'delim' in kw:
            self.delimeter = kw['delim']
            del kw['delim']
        self.comment = r'^#'
        if 'comment' in kw:
            self.comment = kw['comment']
            del kw['comment']
        self.comment = re.compile(self.comment)
        self.outputheaders = False
        if 'outputheaders' in kw:
            self.outputheaders = kw['outputheaders']
            del kw['outputheaders']
        FileTable.__init__(self, *args, **kw)

    def rows(self):
        h = self.open()
        for l in h:
            if self.comment.search(l):
                continue
            r = dict(list(zip(self.headers_ + ["col%d" % i for i in range(len(self.headers_) + 1, 51)],
                         l.split(self.delimeter))))
            yield r
        h.close()

    def from_rows(self, rows):
        wh = self.open(mode='w')
        delim = " "
        if self.delimeter != None:
            delim = self.delimeter
        if self.outputheaders:
            print(delim.join(self.headers_), file=wh)
        for r in rows:
            print(delim.join(map(str, list(map(r.get, self.headers_)))), file=wh)
        wh.close()


class BEDFile(TXTFileTable):
    bedheaders = """
    chrom chromStart chromEnd name score strand thickStart thickEnd itemRgb blockCount blockSizes blockStarts
    """.split()

    def __init__(self, *args, **kw):
        if 'headers' in kw:
            del kw['headers']
        if 'delimiter' in kw:
            del kw['delimiter']
        kw['headers'] = self.bedheaders
        kw['comment'] = r'^(browser|track)\b'
        TXTFileTable.__init__(self, *args, **kw)


class VCFFile(TXTFileTable):

    def __init__(self, *args, **kw):
        kw['headers'] = None
        kw['delim'] = '\t'
        TXTFileTable.__init__(self, *args, **kw)

    def set_headers(self):
        self.infoheaders_ = {}
        self.infoheaders_list_ = []
        self.formatheaders_ = {}
        self.formatheaders_list_ = []
        h = self.open()
        for l in h:
            if l.startswith('##'):
                m = re.search(r'^##INFO=<ID=(\w+),Number=(\w+),Type=(\w+),', l)
                if m:
                    try:
                        n = int(m.group(2))
                    except ValueError:
                        n = m.group(2)
                    self.infoheaders_[m.group(1)] = (n, m.group(3))
                    self.infoheaders_list_.append(m.group(1))
                    continue
                m = re.search(r'^##INFO=<ID=(\w+),', l)
                if m:
                    self.infoheaders_[m.group(1)] = (1, 'String')
                    self.infoheaders_list_.append(m.group(1))
                    continue
                m = re.search(
                    r'^##FORMAT=<ID=(\w+),Number=(\w+),Type=(\w+),', l)
                if m:
                    try:
                        n = int(m.group(2))
                    except ValueError:
                        n = m.group(2)
                    self.formatheaders_[m.group(1)] = (n, m.group(3))
                    self.formatheaders_list_.append(m.group(1))
                    continue
                m = re.search(r'^##FORMAT=<ID=(\w+),', l)
                if m:
                    self.formatheaders_[m.group(1)] = (1, 'String')
                    self.formatheaders_list_.append(m.group(1))
                    continue
                continue
            if l.startswith('#'):
                self.baseheaders_ = l[1:].strip().split('\t')
                break
        try:
            self.fmtpos = self.baseheaders_.index('FORMAT')
        except ValueError:
            self.fmtpos = -1
        if self.fmtpos == -1:
            self.headers_ = self.baseheaders_
            for ih in self.infoheaders_list_:
                n, t = self.infoheaders_[ih]
                if isinstance(n, int) and n > 1 and t in ('Integer', 'Float'):
                    for i in range(1, n + 1):
                        self.headers_.append(ih + ':' + str(i))
                else:
                    self.headers_.append(ih)
        else:
            self.headers_ = self.baseheaders_[:self.fmtpos - 1]
            for ih in self.infoheaders_list_:
                n, t = self.infoheaders_[ih]
                if isinstance(n, int) and n > 1 and t in ('Integer', 'Float'):
                    for i in range(1, n + 1):
                        self.headers_.append(ih + ':' + str(i))
                else:
                    self.headers_.append(ih)
            for fns in self.baseheaders_[self.fmtpos + 1:]:
                self.headers_.extend(
                    [h + ":" + fns for h in list(self.formatheaders_)])
        h.close()

    def rows(self):
        h = self.open()
        for l in h:
            if self.comment.search(l):
                continue
            r = dict(list(zip(self.baseheaders_, l.split(self.delimeter))))
            if 'INFO' in r:
                for kvstr in r['INFO'].split(';'):
                    k, eqstr, vstr = kvstr.partition('=')
                    if not eqstr:
                        vstr = "Y"
                    try:
                        v = str(vstr).strip()
                        v = float(vstr)
                        v = int(vstr)
                    except:
                        pass
                    n, t = self.infoheaders_.get(k,(1,'String'))
                    if isinstance(n, int) and n > 1 and t in ('Integer', 'Float'):
                        if t == 'Integer':
                            vals = list(map(int, v.split(',')))
                        else:
                            vals = list(map(float, v.split(',')))
                        for i in range(len(vals)):
                            r[k + ':' + str(i + 1)] = vals[i]
                        continue
                    r[k] = v
            if 'FORMAT' in r:
                for fns in self.baseheaders_[self.fmtpos + 1:]:
                    for k, vstr in zip(self.formatheaders_list_, r[fns].split(':')):
                        try:
                            v = str(vstr).strip()
                            v = float(vstr)
                            v = int(vstr)
                        except:
                            pass
                        r["%s:%s" % (k, fns)] = v
            yield r
        h.close()

class XLSFileTable(FileTable):

    def __init__(self, *args, **kw):
        if 'sheet' in kw:
            if kw['sheet']:
                self.sheet = kw['sheet']
            else:
                self.sheet = None
            del kw['sheet']
        else:
            self.sheet = None
        self.colwidth = {}
        if 'columnwidth' in kw:
            self.colwidth = kw['columnwidth']
            del kw['columnwidth']
        self.rowheight = {}
        if 'rowheight' in kw:
            self.rowheight = kw['rowheight']
            del kw['rowheight']
        super(XLSFileTable, self).__init__(*args, **kw)

    def set_headers(self):
        import xlrd
        self.headers_ = []
        book = xlrd.open_workbook(self.filename)
        for sheet_idx, sheet_name in enumerate(book.sheet_names()):
            if self.sheet and sheet_name != self.sheet:
                continue
            if not self.sheet and sheet_idx != 0:
                continue
            sheet = book.sheet_by_name(sheet_name)
            for v in sheet.row_values(0):
                self.headers_.append(str(v))
        assert len(self.headers_) > 0

    def rows(self):
        import xlrd
        book = xlrd.open_workbook(self.filename)
        for sheet_idx, sheet_name in enumerate(book.sheet_names()):
            if self.sheet != None and sheet_name != self.sheet:
                continue
            if self.sheet == None and sheet_idx != 0:
                continue
            sheet = book.sheet_by_name(sheet_name)
            for r in range(sheet.nrows):
                if r == 0:
                    continue
                yield dict(list(zip(self.headers_, list(map(str, sheet.row_values(r))))))

    def from_rows(self, rows):
        import xlwt
        wb = xlwt.Workbook()
        assert(self.sheet != None)
        ws = wb.add_sheet(self.sheet)
        dcw = self.colwidth.get(None)
        for j, h in enumerate(self.headers_):
            cw = self.colwidth.get(h, dcw)
            if cw != None:
                XLSFileDataset.setcolwidth(ws, j, cw)
        for j, h in enumerate(self.headers_):
            XLSFileDataset.writevalue(ws, 0, j, h)
        drh = self.rowheight.get(None)
        for i, r in enumerate(rows):
            rh = self.rowheight.get(i, drh)
            if rh:
                XLSFileDataset.setrowheight(ws, i + 1, rh)
            for j, h in enumerate(self.headers_):
                XLSFileDataset.writevalue(ws, i + 1, j, r.get(h, ""))
        wb.save(self.filename)

class XLSXFileTable(FileTable):

    def __init__(self, *args, **kw):
        if 'sheet' in kw:
            if kw['sheet']:
                self.sheet = kw['sheet']
            else:
                self.sheet = None
            del kw['sheet']
        else:
            self.sheet = None
        super(XLSXFileTable, self).__init__(*args, **kw)

    def set_headers(self):
        import openpyxl
        self.headers_ = []
        book = openpyxl.reader.excel.load_workbook(filename=self.filename)
        for sheet_idx, sheet_name in enumerate(book.get_sheet_names()):
            if self.sheet and sheet_name != self.sheet:
                continue
            if not self.sheet and sheet_idx != 0:
                continue
            sheet = book.get_sheet_by_name(sheet_name)
            ncol = sheet.max_column
            for i in range(ncol):
                h = sheet.cell(row=1, column=i+1).value
                if not h:
                    break
                self.headers_.append(str(h))
        assert len(self.headers_) > 0

    def rows(self):
        import openpyxl
        book = openpyxl.reader.excel.load_workbook(filename=self.filename)
        for sheet_idx, sheet_name in enumerate(book.get_sheet_names()):
            if self.sheet != None and sheet_name != self.sheet:
                continue
            if self.sheet == None and sheet_idx != 0:
                continue
            sheet = book.get_sheet_by_name(sheet_name)
            ncol = sheet.max_column
            nrow = sheet.max_row
            for r in range(nrow):
                if r == 0:
                    continue
                row = []
                for c in range(ncol):
                    value = sheet.cell(row=r+1, column=c+1).value
                    row.append(value)
                yield dict(list(zip(self.headers_, row)))

    def from_rows(self, rows):
        import openpyxl
        wb = openpyxl.workbook.Workbook()
        assert(self.sheet != None)
        ws = wb.worksheets[0]
        ws.title = self.sheet
        # ws.auto_filter = 'A1:%s1' % (
        #     openpyxl.utils.cell.get_column_letter(len(self.headers_)),)
        for j, h in enumerate(self.headers_):
            ws.cell(row=1, column=j+1).value = h
        for i, r in enumerate(rows):
            for j, h in enumerate(self.headers_):
                XLSXFileDataset.writevalue(ws, i + 2, j+1, r.get(h))
        wb.save(filename=self.filename)


class AddField(Table):

    def __init__(self, tablein, fieldname, valuefunction):
        self.tin = tablein
        self.fieldname = fieldname
        self.headers_ = list(tablein.headers())
        self.headers_.append(fieldname)
        self.f = valuefunction

    def rows(self):
        for r in self.tin:
            r[self.fieldname] = self.f(r)
            yield r


class SetField(Table):

    def __init__(self, tablein, fieldname, valuefunction):
        self.tin = tablein
        self.headers_ = list(tablein.headers())
        self.fieldname = fieldname
        self.f = valuefunction

    def rows(self):
        for r in self.tin:
            r[self.fieldname] = self.f(r)
            yield r


class AddFields(Table):

    def __init__(self, tablein, fieldfunction, valuefunction):
        self.tin = tablein
        self.headers_ = fieldfunction(tablein.headers())
        self.f = valuefunction

    def rows(self):
        for r in self.tin:
            yield self.f(r)


class ReplaceFromList(Table):

    def __init__(self, tablein, field, values):
        self.tin = tablein
        self.headers_ = tablein.headers()
        assert field in self.headers_
        self.f = field
        self.l = values

    def rows(self):
        for r, v in zip(self.tin, self.l):
            r[self.f] = v
            yield r


class FilterRowsByFunction(Table):

    def __init__(self, tablein, fn):
        self.tin = tablein
        self.headers_ = tablein.headers()
        self.f = fn

    def rows(self):
        for r in self.tin:
            if self.f(r):
                yield r


class FilterRows(Table):

    def __init__(self, tablein, expr):
        self.tin = tablein
        self.headers_ = tablein.headers()
        self.expr = expr
        self.varmap = {}
        self.listmap = {}
        j = 0
        for h in sorted(self.headers_, key=lambda h: -len(h)):
            if h in self.expr:
                var = "v%d" % j
                self.expr = self.expr.replace(h, var)
                self.varmap[h] = var
                j += 1
        for h in sorted(self.headers_, key=lambda h: -len(h)):
            m = re.search(r'^(.*?)\d+$', h)
            if m and m.group(1) in self.expr:
                var = "v%d" % j
                self.expr = self.expr.replace(m.group(1), var)
                self.listmap[m.group(1)] = var
                j += 1

    def rows(self):
        for r in self.tin:
            locals = {}
            for h in self.listmap:
                locals[self.listmap[h]] = []
                any = False
                for i in range(21):
                    try:
                        locals[self.listmap[h]].append(
                            float(r["%s%d" % (h, i)]))
                        any = True
                    except KeyError:
                        if not any:
                            locals[self.listmap[h]].append(None)
                        else:
                            break
            for h in self.varmap:
                try:
                    locals[self.varmap[h]] = float(r[h])
                except ValueError:
                    locals[self.varmap[h]] = r[h]
            # print >>sys.stderr, self.expr,locals
            if eval(self.expr, globals(), locals):
                yield r


class HeaderMap(Table):

    def __init__(self, tablein, headermap):
        self.tin = tablein
        self.headers_ = list(map(headermap, tablein.headers()))
        self.headermap = headermap

    def rows(self):
        for r in self.tin:
            yield dict([(self.headermap(k), v) for (k, v) in list(r.items())])


class ColumnSelect(Table):

    def __init__(self, tablein, goodset):
        self.tin = tablein
        self.headers_ = [h for h in tablein.headers() if h in goodset]
        self.goodset = goodset

    def rows(self):
        for r in self.tin:
            yield dict([(k, v) for (k, v) in list(r.items()) if k in self.goodset])


class ColumnRemove(Table):

    def __init__(self, tablein, badset):
        self.tin = tablein
        self.headers_ = [h for h in tablein.headers() if h not in badset]
        self.badset = badset

    def rows(self):
        for r in self.tin:
            yield dict([(k, v) for (k, v) in list(r.items()) if k not in self.badset])


class ColumnRemoveRegex(Table):

    def __init__(self, tablein, badregex):
        self.tin = tablein
        self.headers_ = [
            h for h in tablein.headers() if not re.search(badregex, h)]

    def rows(self):
        for r in self.tin:
            yield dict([(k, v) for (k, v) in list(r.items()) if k in self.headers_])


class ValueMap(Table):

    def __init__(self, tablein, mapfn, colset=None):
        self.tin = tablein
        self.headers_ = list(tablein.headers())
        self.f = mapfn
        self.colset = colset

    def applymap(self, item):
        if self.colset == None or item[0] in self.colset:
            return item[0], self.f(item[1])
        return item

    def rows(self):
        for r in self.tin:
            yield dict(list(map(self.applymap, list(r.items()))))


class MoveField(Table):

    def __init__(self, tablein, pos, cols, after=True):
        self.tin = tablein
        self.headers_ = list(tablein.headers())
        for c in cols:
            assert c in self.headers_
        try:
            p = int(pos)
        except (TypeError, ValueError):
            assert pos in self.headers_
            p = self.headers_.index(pos)
        for c in reversed(cols):
            p0 = self.headers_.index(c)
            del self.headers_[p0]
            if after:
                self.headers_.insert(p + 1, c)
            else:
                self.headers_.insert(p, c)

    def rows(self):
        for r in self.tin:
            yield r


class ParsimonyCSV:

    def __init__(self, proteins):
        self._proteins = set(proteins)

    def rows(self, inrows):
        for r in inrows:
            proteins = set(r['protein'].split(';'))
            proteins &= self._proteins
            if len(proteins) > 0:
                r['protein'] = ';'.join(proteins)
                yield r

    def rewrite(self, infile, outfile):
        rows = CSVFileTable(filename=infile)
        newrows = CSVFileTable(
            filename=outfile, headers=rows.headers(), compression=None)
        newrows.from_rows(self.rows(rows))
